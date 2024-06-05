import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import sys
import time
import tkinter as tk
from tkinter import filedialog, messagebox
import threading

def remove_null_columns(filename):
    start_time = time.time()  # Record start time

    # Check file type and read into DataFrame
    if filename.endswith('.csv'):
        df = pd.read_csv(filename)
    elif filename.endswith('.xlsx'):
        df = pd.read_excel(filename)
    else:
        #print("Unsupported file format. Please provide a CSV or XLSX file.")
        messagebox.showerror("Error", f"Unsupported file format for '{filename}'. Please provide a CSV or XLSX file.")
        return

    # Check if the number of columns is more than 300
    if len(df.columns) > 300:
        print("The file has more than 300 columns. Exiting the program.")
        return

    # Find 'eventTime' column and move it to the first column
    if 'eventTime' in df.columns:
        cols = list(df.columns)
        cols.insert(0, cols.pop(cols.index('eventTime'))) 
        df = df[cols]

    # Identify columns with only 'null' or NaN values and delete them, except 'productAction'
    columns_to_delete = []
    for col in df.columns:
        # Replace NaN with 'null' for comparison purposes
        df[col] = df[col].fillna('null')
        unique_values = df[col].unique()

        if len(unique_values) == 1 and str(unique_values[0]).strip().lower() == 'null':
            if col != 'productAction':
                columns_to_delete.append(col)

    # Debugging output
    #print("Columns to delete:", columns_to_delete)

    # Drop identified columns
    df = df.drop(columns=columns_to_delete)

    # Additional processing for columns starting with 'aisaac' and ending with 'Name'
    i = 0
    while i < len(df.columns) - 1:
        current_col = df.columns[i]
        next_col = df.columns[i + 1]

        if current_col.startswith('aisaac') and current_col.endswith('Name'):
            if next_col + 'Name' != current_col:
                df = df.drop(columns=[current_col])
            else:
                if df[current_col].notna().all() and df[current_col].nunique() == 1 and str(df[current_col].iloc[0]).strip().lower() != 'null':
                    unique_value = df[current_col].iloc[0]
                    df.rename(columns={next_col: unique_value}, inplace=True)
                    df = df.drop(columns=[current_col])
                else:
                    i += 1
        else:
            i += 1

    # Columns to delete
    columns_to_delete = [
        'rawEvent', 'rawEventHash', 'cenNifiSentTime', 'customerURI',
        'cenNifiReceiptTime', 'logFilterKafkaInTime', 'logFilterInTime'
    ]
    customer_uri_value = df['customerURI'].iloc[0] if 'customerURI' in df.columns else 'output'
    df = df.drop(columns=[col for col in columns_to_delete if col in df.columns])

    # Get customerURI value for the output filename
    #customer_uri_value = df['customerURI'].iloc[0] if 'customerURI' in df.columns else 'output'
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_filename = f"{customer_uri_value}_{timestamp}.xlsx"

    # Create an Excel workbook and add the DataFrame
    wb = openpyxl.Workbook()
    ws = wb.active

    # Add DataFrame to worksheet, keeping 'null' as 'null'
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Style formatting for the header row
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=1), start=1):
        for cell in row:
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                 right=openpyxl.styles.Side(style='thin'),
                                                 top=openpyxl.styles.Side(style='thin'),
                                                 bottom=openpyxl.styles.Side(style='thin'))
            cell.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Style formatting for data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                 right=openpyxl.styles.Side(style='thin'),
                                                 top=openpyxl.styles.Side(style='thin'),
                                                 bottom=openpyxl.styles.Side(style='thin'))
            cell.font = Font(name='Calibri', size=12, color='000000')

    # Adjust column width based on content
    for column_cells in ws.columns:
        max_length = 0
        column = openpyxl.utils.get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.1
        if adjusted_width > 40:
            adjusted_width = 40
        ws.column_dimensions[column].width = adjusted_width

    # Save the modified workbook
    try:
        wb.save(output_filename)
        elapsed_time = time.time() - start_time  # Calculate elapsed time
        messagebox.showinfo("Success", f"Modified workbook saved as '{output_filename}'.\nElapsed Time: {elapsed_time:.2f} seconds")
    except PermissionError:
        messagebox.showerror("Error", "Permission denied. Please make sure you have write access to the directory.")
        return

def remove_null_columns_threaded(filename):
    threading.Thread(target=remove_null_columns, args=(filename,), daemon=True).start()

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("CSV and Excel files", "*.csv *.xlsx"), ("CSV files", "*.csv"), ("Excel files", "*.xlsx")])
    if file_path:
        remove_null_columns_threaded(file_path)

def exit_application():
    sys.exit()

# Create GUI window with a beige background color similar to the image.
root = tk.Tk()
root.title("Normalizer")
root.configure(bg='#f5f5dc')  # Beige background color

# Create a label with a grey background and white text for the title.
title_label = tk.Label(root, text="Normalizer", bg='#808080', fg='white', font=('Arial', 24), width=30)
title_label.pack(padx=20, pady=(20, 0))

# Update buttons with grey background and white text.
select_button = tk.Button(root, text="Select File", command=select_file, bg='#808080', fg='white', font=('Arial', 14), width=15)
select_button.pack(pady=(40, 10))

exit_button = tk.Button(root, text="Exit", command=exit_application, bg='#808080', fg='white', font=('Arial', 14), width=15)
exit_button.pack(pady=(0, 40))

# Align the developer's name in the bottom right-most corner.
developer_label = tk.Label(root, text="G.S", bg='#f5f5dc', fg='black', font=('Arial', 10))
developer_label.pack(side=tk.BOTTOM, anchor='e', padx=10, pady=(0, 10))

root.mainloop()
