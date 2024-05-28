import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import sys
import os
import time
import tkinter as tk
from tkinter import filedialog, messagebox
import threading

def remove_null_columns(filename):
    start_time = time.time()  # Record start time
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        messagebox.showerror("Error", f"The specified Excel file '{filename}' was not found.")
        return

    sheet = wb.active
    columns_to_delete = []
    column_positions = {cell.value: cell.column for cell in sheet[1]}

    # Identify columns to delete
    delete_columns_names = ['rawEvent', 'rawEventHash', 'aisaacReceivedTime', 'customerURI', 'cenNifiReceiptTime', 'logFilterKafkaInTime', 'logFilterInTime']
    columns_to_delete.extend([col_idx for col_name, col_idx in column_positions.items() if col_name in delete_columns_names])

    # Move 'eventTime' to the first column
    if 'eventTime' in column_positions:
        event_time_col_idx = column_positions['eventTime']
        sheet.insert_cols(1)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            sheet.cell(row=row[0].row, column=1, value=sheet.cell(row=row[0].row, column=event_time_col_idx + 1).value)
        sheet.delete_cols(event_time_col_idx + 1)
        # Adjust the positions after moving 'eventTime'
        for col_name in column_positions:
            if column_positions[col_name] > event_time_col_idx:
                column_positions[col_name] -= 1

    # Identify columns with names ending in "Name" and process them
    name_columns = [col_idx for col_name, col_idx in column_positions.items()
                    if col_name and col_name.endswith("Name")]

    for col_idx in name_columns:
        unique_values = set()
        for row_idx in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value and str(cell_value).strip().lower() != "null":
                unique_values.add(cell_value)

        if len(unique_values) == 1:
            unique_value = unique_values.pop()
            adjacent_col_idx = col_idx + 1
            if adjacent_col_idx <= sheet.max_column:
                header_value = sheet.cell(row=1, column=col_idx).value
                adjacent_header_value = sheet.cell(row=1, column=adjacent_col_idx).value
                if adjacent_header_value and header_value.replace("Name", "") == adjacent_header_value:
                    sheet.cell(row=1, column=adjacent_col_idx).value = unique_value
                    columns_to_delete.append(col_idx)

    # Iterate over each column to find null columns
    for col_idx in range(sheet.max_column, 0, -1):  # Start from the last column
        if col_idx in columns_to_delete:
            continue
        is_null_column = True  # Assume the column is full of "null" until proven otherwise

        for row_idx in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            # Check if the cell is not "null" (case insensitive)
            if cell_value is not None and str(cell_value).strip().lower() != "null":
                is_null_column = False
                break  # No need to check further; this column is not full of "null"

        if is_null_column:
            columns_to_delete.append(col_idx)

    # Delete identified columns
    for col_index in sorted(columns_to_delete, reverse=True):
        sheet.delete_cols(col_index)

    # Style formatting for the header row
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=1), start=1):
        for cell in row:
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                 right=openpyxl.styles.Side(style='thin'),
                                                 top=openpyxl.styles.Side(style='thin'),
                                                 bottom=openpyxl.styles.Side(style='thin'))
            cell.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Style formatting for data rows
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                 right=openpyxl.styles.Side(style='thin'),
                                                 top=openpyxl.styles.Side(style='thin'),
                                                 bottom=openpyxl.styles.Side(style='thin'))
            cell.font = Font(name='Calibri', size=12, color='000000')

    # Adjust column width based on content
    for column_cells in sheet.columns:
        max_length = 0
        column = openpyxl.utils.get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    # Save the modified workbook
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_filename = filename.replace('.xlsx', f'_{timestamp}.xlsx')
    try:
        wb.save(output_filename)
        elapsed_time = time.time() - start_time  # Calculate elapsed time
        messagebox.showinfo("Success", f"Modified workbook saved as '{output_filename}'.\nElapsed Time: {elapsed_time:.2f} seconds")
    except PermissionError:
        messagebox.showerror("Error", "Permission denied. Please make sure you have write access to the directory.")

def remove_null_columns_threaded(filename):
    threading.Thread(target=remove_null_columns, args=(filename,), daemon=True).start()

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        remove_null_columns_threaded(file_path)

def exit_application():
    sys.exit()

# Create GUI window
root = tk.Tk()
root.title("Excel Column Remover")

# Create buttons
select_button = tk.Button(root, text="Select File", command=select_file)
select_button.pack(pady=10)

exit_button = tk.Button(root, text="Exit", command=exit_application)
exit_button.pack(pady=5)

root.mainloop()
